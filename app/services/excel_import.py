from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook

from app.models import Transaction
<<<<<<< HEAD
from app.utils.constants import TransactionKind
=======
from app.utils.constants import TransactionType
>>>>>>> origin/main


def parse_transactions(path: Path) -> List[Transaction]:
    wb = load_workbook(path)
<<<<<<< HEAD
    parsed: list[Transaction] = []

    if "Повседневные" in wb.sheetnames:
        ws = wb["Повседневные"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            parsed.append(
                Transaction(
                    date=row[2],
                    title=row[3],
                    amount=row[6],
                    comment=row[7],
                    kind=TransactionKind.DAILY,
                )
            )
    if "Крупные" in wb.sheetnames:
        ws = wb["Крупные"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            parsed.append(
                Transaction(
                    date=row[1],
                    title=row[2],
                    amount=row[3],
                    kind=TransactionKind.BIG,
                )
            )
    if "Квартира" in wb.sheetnames:
        ws = wb["Квартира"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            parsed.append(
                Transaction(
                    date=row[1],
                    title=row[3],
                    amount=row[4],
                    kind=TransactionKind.HOME,
                )
            )
=======
    mapping = {
        "Повседневные": TransactionType.DAILY,
        "Крупные": TransactionType.BIG,
        "Квартира": TransactionType.APARTMENT,
    }
    parsed: list[Transaction] = []
    for sheet_name, tx_type in mapping.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            tx = Transaction(
                tx_date=row[2],
                title=row[3],
                amount=row[6],
                comment=row[7],
                type=tx_type,
            )
            parsed.append(tx)
>>>>>>> origin/main
    return parsed


def deduplicate(
    existing: Iterable[Transaction], imported: Iterable[Transaction]
) -> List[Transaction]:
    existing_keys = {
<<<<<<< HEAD
        (
            tx.kind,
            tx.date,
            (tx.title or "").strip().lower(),
            float(tx.amount),
            getattr(tx.category, "name", None),
            getattr(tx.subcategory, "name", None),
        )
=======
        (tx.type, tx.tx_date, (tx.title or "").strip().lower(), float(tx.amount))
>>>>>>> origin/main
        for tx in existing
    }
    result: list[Transaction] = []
    for tx in imported:
<<<<<<< HEAD
        key = (
            tx.kind,
            tx.date,
            (tx.title or "").strip().lower(),
            float(tx.amount),
            None,
            None,
        )
=======
        key = (tx.type, tx.tx_date, (tx.title or "").strip().lower(), float(tx.amount))
>>>>>>> origin/main
        if key in existing_keys:
            continue
        result.append(tx)
    return result
