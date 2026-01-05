<<<<<<< HEAD
from calendar import day_abbr
from pathlib import Path
from typing import Iterable, Sequence
=======
from pathlib import Path
from typing import Iterable
>>>>>>> origin/main

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Category, Subcategory, Transaction
<<<<<<< HEAD
from app.utils.constants import TransactionKind
=======
from app.utils.constants import TransactionType
>>>>>>> origin/main

TEMPLATE_FILE = Path("Копия Costs.xlsx")


<<<<<<< HEAD
def _write_daily(ws: Worksheet, transactions: Sequence[Transaction]) -> None:
    row = 2
    for tx in transactions:
        ws.cell(row=row, column=1, value=day_abbr[tx.date.weekday()])
        ws.cell(row=row, column=2, value=tx.date.strftime("%Y-%m"))
        ws.cell(row=row, column=3, value=tx.date)
=======
def _write_transactions(ws: Worksheet, transactions: Iterable[Transaction]) -> None:
    row = 2
    for tx in transactions:
        ws.cell(row=row, column=1, value=tx.tx_date.strftime("%a"))
        ws.cell(row=row, column=2, value=tx.tx_date.strftime("%Y-%m"))
        ws.cell(row=row, column=3, value=tx.tx_date)
>>>>>>> origin/main
        ws.cell(row=row, column=4, value=tx.title or "")
        ws.cell(row=row, column=5, value=tx.category.name if tx.category else "")
        ws.cell(row=row, column=6, value=tx.subcategory.name if tx.subcategory else "")
        ws.cell(row=row, column=7, value=float(tx.amount))
        ws.cell(row=row, column=8, value=tx.comment or "")
        row += 1


<<<<<<< HEAD
def _write_big(ws: Worksheet, transactions: Sequence[Transaction]) -> None:
    row = 2
    for tx in transactions:
        ws.cell(row=row, column=1, value=tx.date.strftime("%Y-%m"))
        ws.cell(row=row, column=2, value=tx.date)
        ws.cell(row=row, column=3, value=tx.title or "")
        ws.cell(row=row, column=4, value=float(tx.amount))
        row += 1


def _write_home(ws: Worksheet, transactions: Sequence[Transaction]) -> None:
    row = 2
    for tx in transactions:
        ws.cell(row=row, column=1, value=tx.date.strftime("%Y-%m"))
        ws.cell(row=row, column=2, value=tx.date)
        ws.cell(row=row, column=3, value=tx.category.name if tx.category else "")
        ws.cell(row=row, column=4, value=tx.title or "")
        ws.cell(row=row, column=5, value=float(tx.amount))
        row += 1


=======
>>>>>>> origin/main
def export_to_xlsx(
    *,
    transactions: Iterable[Transaction],
    categories: Iterable[Category],
    subcategories: Iterable[Subcategory],
    output_path: Path,
) -> Path:
    if TEMPLATE_FILE.exists():
        wb = load_workbook(TEMPLATE_FILE)
    else:
        wb = Workbook()
<<<<<<< HEAD
    tx_by_kind: dict[TransactionKind, list[Transaction]] = {
        TransactionKind.DAILY: [],
        TransactionKind.BIG: [],
        TransactionKind.HOME: [],
    }
    for tx in transactions:
        tx_by_kind[tx.kind].append(tx)

    daily_sheet = wb["Повседневные"] if "Повседневные" in wb.sheetnames else wb.create_sheet("Повседневные")
    _write_daily(daily_sheet, tx_by_kind[TransactionKind.DAILY])

    big_sheet = wb["Крупные"] if "Крупные" in wb.sheetnames else wb.create_sheet("Крупные")
    _write_big(big_sheet, tx_by_kind[TransactionKind.BIG])

    home_sheet = wb["Квартира"] if "Квартира" in wb.sheetnames else wb.create_sheet("Квартира")
    _write_home(home_sheet, tx_by_kind[TransactionKind.HOME])

    ref_sheet = wb["Справочники"] if "Справочники" in wb.sheetnames else wb.create_sheet("Справочники")
    ref_sheet.cell(row=1, column=1, value="Категория")
    ref_sheet.cell(row=1, column=2, value="Подкатегория")
    idx = 2
    for cat in categories:
        ref_sheet.cell(row=idx, column=1, value=cat.name)
        idx += 1
        for sub in [s for s in subcategories if s.category_id == cat.id]:
            ref_sheet.cell(row=idx, column=1, value=cat.name)
            ref_sheet.cell(row=idx, column=2, value=sub.name)
            idx += 1
=======
    sheet_mapping = {
        TransactionType.DAILY: "Повседневные",
        TransactionType.BIG: "Крупные",
        TransactionType.APARTMENT: "Квартира",
    }
    tx_by_type: dict[TransactionType, list[Transaction]] = {
        TransactionType.DAILY: [],
        TransactionType.BIG: [],
        TransactionType.APARTMENT: [],
    }
    for tx in transactions:
        tx_by_type[tx.type].append(tx)

    for tx_type, sheet_name in sheet_mapping.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        _write_transactions(ws, tx_by_type[tx_type])

    if "Справочники" in wb.sheetnames:
        ws_ref = wb["Справочники"]
    else:
        ws_ref = wb.create_sheet("Справочники")
    ws_ref.cell(row=1, column=1, value="Категория")
    ws_ref.cell(row=1, column=2, value="Подкатегория")
    idx = 2
    for cat in categories:
        ws_ref.cell(row=idx, column=1, value=cat.name)
        idx += 1
        for sub in subcategories:
            if sub.category_id == cat.id:
                ws_ref.cell(row=idx, column=1, value=cat.name)
                ws_ref.cell(row=idx, column=2, value=sub.name)
                idx += 1
>>>>>>> origin/main

    wb.save(output_path)
    return output_path
